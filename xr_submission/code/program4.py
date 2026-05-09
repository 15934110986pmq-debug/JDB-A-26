# -*- coding: utf-8 -*-
"""
第四问第二版代码：候选执行窗—冲突图—整数规划任务选择

适配第二版论文写法：
1. 从第三问 10Hz 轨迹出发，构造高密度扫描轨迹；
2. 对射击/拍照目标生成严格可执行候选窗；
3. 射击目标最多执行一次；
4. 拍照目标允许多次，但同一目标任意两次方向角差 >= 60°；
5. 通过冲突图 + 0-1整数规划最大化任务数量；
6. 严格检查距离、速度、加速度、准备时间、时间冲突、角度差；
7. 保护 result.xlsx 中红色文字单元格。

运行：
    cd /d E:\数模A
    python program4.py
"""

from __future__ import annotations

import math
import shutil
import warnings
from dataclasses import dataclass
from pathlib import Path
from collections import defaultdict

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy.interpolate import PchipInterpolator
from scipy.signal import savgol_filter
from scipy.sparse import coo_matrix
from scipy.optimize import milp, LinearConstraint, Bounds
from openpyxl import load_workbook, Workbook


# ============================================================
# 0. 参数设置
# ============================================================

BASE_DIR = Path.cwd()
OUT_DIR = BASE_DIR / "q4_second_outputs"

Q3_DIR_CANDIDATES = [
    # 第四问应优先使用第三问为任务规划准备的完整/扩展轨迹，
    # 避免误读 q3_outputs 中较短的公共区间轨迹导致任务数偏低。
    BASE_DIR / "q3_final_outputs",
    BASE_DIR / "q3_outputs",
    BASE_DIR,
]

TRAJ_FILE_NAMES = [
    "q3_extended_10hz_trajectory.csv",
    "q3_for_q4_10hz_trajectory.csv",
    "q3_final_10hz_trajectory.csv",
    "q3_fused_10hz_trajectory.csv",
    "q3_fusion_10hz_trajectory.csv",
    "q3_10hz_trajectory.csv",
]

TARGET_FILE_CANDIDATES = [
    BASE_DIR / "附件4.xlsx",
    BASE_DIR / "附件4_目标点.xlsx",
    BASE_DIR / "附件4：目标点.xlsx",
    BASE_DIR / "targets.xlsx",
]

RESULT_TEMPLATE_CANDIDATES = [
    BASE_DIR / "result.xlsx",
    BASE_DIR / "结果模板.xlsx",
]

# 轨迹扫描频率
SCAN_HZ = 300.0
SCAN_DT = 1.0 / SCAN_HZ

# 任务硬约束
SHOOT_D_MIN = 5.0
SHOOT_D_MAX = 30.0
SHOOT_V_MAX = 2.0
SHOOT_A_MAX = 1.5
SHOOT_PREP = 1.5
SHOOT_HIT_PROB = 0.85

PHOTO_D_MIN = 10.0
PHOTO_D_MAX = 40.0
PHOTO_V_MAX = 1.5
PHOTO_A_MAX = 1.5
PHOTO_PREP = 0.5
PHOTO_ANGLE_MIN = 60.0

# 候选压缩参数
SHOOT_STEP_S = 0.03
PHOTO_STEP_S = 0.03
SHOOT_MAX_CAND_PER_TARGET = 220
PHOTO_MAX_CAND_PER_TARGET = 520
PHOTO_ANGLE_BIN = 5.0

# MILP 参数
MILP_TIME_LIMIT = 1200
LOCAL_POOL_SIZE = 1800
LOCAL_SWAP_POOL_SIZE = 260

# 目标函数权重。总任务数是第一优先级。
W_TOTAL = 1_000_000_000.0
W_PHOTO_COVER = 10_000_000.0
W_SHOOT_COUNT = 2_000_000.0
W_QUALITY = 1_000.0
W_EARLY = 0.001


# ============================================================
# 1. 数据结构
# ============================================================

@dataclass(frozen=True)
class TaskWindow:
    uid: int
    target_id: str
    task_type: str              # "射击" / "拍照"
    prep_start: float
    execute_time: float
    occupy_end: float
    robot_x: float
    robot_y: float
    target_x: float
    target_y: float
    distance: float
    speed: float
    accel: float
    view_angle: float
    quality: float


# ============================================================
# 2. 通用工具
# ============================================================

def ensure_output_dir() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)


def norm_text(x: object) -> str:
    if x is None:
        return ""
    return (
        str(x)
        .strip()
        .replace(" ", "")
        .replace("\n", "")
        .replace("\r", "")
        .replace("（", "(")
        .replace("）", ")")
        .lower()
    )


def pick_col(df: pd.DataFrame, aliases: list[str], required: bool = True) -> str | None:
    col_map = {norm_text(c): c for c in df.columns}
    alias_norm = [norm_text(a) for a in aliases]

    for a in alias_norm:
        if a in col_map:
            return col_map[a]

    for key, raw in col_map.items():
        for a in alias_norm:
            if a in key or key in a:
                return raw

    if required:
        raise KeyError(f"找不到列：{aliases}；当前列为：{list(df.columns)}")
    return None


def find_first_existing(candidates: list[Path], desc: str) -> Path:
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError(f"未找到{desc}，候选路径：\n" + "\n".join(map(str, candidates)))


def read_csv_auto(path: Path) -> pd.DataFrame:
    for enc in ["utf-8-sig", "utf-8", "gbk", "gb2312"]:
        try:
            return pd.read_csv(path, encoding=enc)
        except Exception:
            pass
    raise RuntimeError(f"无法读取CSV文件：{path}")


def angle_diff_deg(a: float, b: float) -> float:
    d = abs(a - b) % 360.0
    return min(d, 360.0 - d)


def bearing_deg(rx: float, ry: float, tx: float, ty: float) -> float:
    return math.degrees(math.atan2(ty - ry, tx - rx)) % 360.0


def continuous_true_blocks(mask: np.ndarray) -> list[tuple[int, int]]:
    mask = np.asarray(mask, dtype=bool)
    if mask.size == 0:
        return []
    diff = np.diff(np.r_[False, mask, False].astype(int))
    starts = np.where(diff == 1)[0]
    ends = np.where(diff == -1)[0] - 1
    return list(zip(starts, ends))


def all_true_in_previous_window(mask: np.ndarray, n_window: int) -> np.ndarray:
    """ready[i] 表示 i 时刻及之前 n_window 个采样点全部满足约束。"""
    mask = np.asarray(mask, dtype=bool)
    n = len(mask)
    ready = np.zeros(n, dtype=bool)

    if n_window <= 1:
        return mask.copy()
    if n < n_window:
        return ready

    cs = np.r_[0, np.cumsum(mask.astype(int))]
    cnt = cs[n_window:] - cs[:-n_window]
    ready[n_window - 1:] = cnt == n_window
    return ready


def smooth_array(arr: np.ndarray, window: int = 31, poly: int = 3) -> np.ndarray:
    arr = np.asarray(arr, dtype=float)
    n = len(arr)
    if n < 7:
        return arr.copy()

    w = min(window, n)
    if w % 2 == 0:
        w -= 1
    if w <= poly + 1:
        w = poly + 3
        if w % 2 == 0:
            w += 1
    if w > n:
        w = n if n % 2 == 1 else n - 1
    if w < 5:
        return arr.copy()

    return savgol_filter(arr, window_length=w, polyorder=min(poly, w - 2), mode="interp")


def circular_span_deg(angles: np.ndarray) -> float:
    angles = np.asarray(angles, dtype=float)
    if len(angles) == 0:
        return np.nan
    if len(angles) == 1:
        return 0.0
    a = np.sort(angles % 360.0)
    gaps = np.diff(np.r_[a, a[0] + 360.0])
    return float(360.0 - np.max(gaps))


# ============================================================
# 3. 读取轨迹和目标
# ============================================================

def find_q3_trajectory() -> Path:
    """
    优先选择第三问为第四问准备的完整/扩展轨迹。
    若多个轨迹文件同时存在，则优先级为：
    1. q3_final_outputs 目录；
    2. 文件名含 extended / q4 / final；
    3. 行数更多的轨迹文件；
    4. CSV 文件。
    """
    exact_pool: list[Path] = []

    for d in Q3_DIR_CANDIDATES:
        for name in TRAJ_FILE_NAMES:
            p = d / name
            if p.exists():
                exact_pool.append(p)

    pool: list[Path] = list(exact_pool)

    for d in Q3_DIR_CANDIDATES:
        if not d.exists():
            continue
        for p in d.glob("*"):
            if p.suffix.lower() not in [".csv", ".xlsx", ".xls"]:
                continue
            if p in pool:
                continue

            name = norm_text(p.name)
            bad = ["summary", "diagnosis", "bootstrap", "residual", "残差", "诊断", "统计", "result"]
            good = ["trajectory", "traj", "轨迹", "10hz", "extended", "扩展", "q4", "融合", "fused", "final"]

            if any(b in name for b in bad):
                continue
            if any(g in name for g in good):
                pool.append(p)

    if not pool:
        raise FileNotFoundError("未找到第三问输出的10Hz轨迹文件。")

    def quick_row_count(p: Path) -> int:
        try:
            if p.suffix.lower() == ".csv":
                return max(0, sum(1 for _ in open(p, "rb")) - 1)
            df_head = pd.read_excel(p)
            return len(df_head)
        except Exception:
            return 0

    def score(p: Path) -> tuple[int, int, int, int, int, int]:
        name = norm_text(p.name)
        parent = norm_text(p.parent.name)
        return (
            1 if "q3_final_outputs" in parent else 0,
            1 if ("extended" in name or "扩展" in name or "q4" in name or "final" in name) else 0,
            1 if "10hz" in name else 0,
            1 if ("trajectory" in name or "traj" in name or "轨迹" in name) else 0,
            quick_row_count(p),
            1 if p.suffix.lower() == ".csv" else 0,
        )

    return sorted(pool, key=score, reverse=True)[0]


def read_trajectory(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        df = read_csv_auto(path)
    else:
        df = pd.read_excel(path)

    t_col = pick_col(df, ["time", "t", "时间", "时刻", "time_s", "t_s"])
    x_col = pick_col(df, ["x", "x_m", "x坐标", "x坐标(m)", "X坐标(m)", "X"])
    y_col = pick_col(df, ["y", "y_m", "y坐标", "y坐标(m)", "Y坐标(m)", "Y"])

    out = pd.DataFrame({
        "time": pd.to_numeric(df[t_col], errors="coerce"),
        "x": pd.to_numeric(df[x_col], errors="coerce"),
        "y": pd.to_numeric(df[y_col], errors="coerce"),
    })

    out = out.dropna(subset=["time", "x", "y"])
    out = out.sort_values("time").drop_duplicates(subset=["time"], keep="first").reset_index(drop=True)

    if len(out) < 10:
        raise ValueError("第三问轨迹点数过少。")

    return out


def densify_trajectory(df: pd.DataFrame) -> pd.DataFrame:
    """10Hz轨迹平滑后加密到300Hz，并计算平滑速度/加速度。"""
    t = df["time"].to_numpy(float)
    x = smooth_array(df["x"].to_numpy(float), window=21, poly=3)
    y = smooth_array(df["y"].to_numpy(float), window=21, poly=3)

    t0 = math.ceil(float(t.min()) * SCAN_HZ) / SCAN_HZ
    t1 = math.floor(float(t.max()) * SCAN_HZ) / SCAN_HZ
    if t1 <= t0:
        raise ValueError("轨迹时间范围异常，无法加密。")

    t_new = np.arange(t0, t1 + 0.5 * SCAN_DT, SCAN_DT)

    fx = PchipInterpolator(t, x, extrapolate=False)
    fy = PchipInterpolator(t, y, extrapolate=False)

    x_new = fx(t_new)
    y_new = fy(t_new)

    x_new = smooth_array(x_new, window=61, poly=3)
    y_new = smooth_array(y_new, window=61, poly=3)

    vx = np.gradient(x_new, SCAN_DT)
    vy = np.gradient(y_new, SCAN_DT)
    speed = np.sqrt(vx * vx + vy * vy)
    speed = smooth_array(speed, window=61, poly=3)

    accel = np.abs(np.gradient(speed, SCAN_DT))
    accel = smooth_array(accel, window=61, poly=3)

    dense = pd.DataFrame({
        "time": t_new,
        "x": x_new,
        "y": y_new,
        "speed": speed,
        "accel": accel,
    })

    dense = dense.replace([np.inf, -np.inf], np.nan).dropna().reset_index(drop=True)
    return dense


def read_targets(path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    sheets = pd.read_excel(path, sheet_name=None)

    shoot_sheet = None
    photo_sheet = None

    for name in sheets:
        n = norm_text(name)
        if "射击" in n or "shoot" in n:
            shoot_sheet = name
        if "拍照" in n or "photo" in n or "camera" in n:
            photo_sheet = name

    if shoot_sheet is None or photo_sheet is None:
        names = list(sheets.keys())
        if len(names) < 2:
            raise ValueError("附件4至少应包含射击目标和拍照目标两个工作表。")
        shoot_sheet, photo_sheet = names[0], names[1]

    def clean(raw: pd.DataFrame, kind: str) -> pd.DataFrame:
        id_col = pick_col(raw, ["编号", "目标编号", "id", "target_id"])
        x_col = pick_col(raw, ["x坐标(m)", "x坐标", "x", "X坐标(m)", "X"])
        y_col = pick_col(raw, ["y坐标(m)", "y坐标", "y", "Y坐标(m)", "Y"])

        temp = pd.DataFrame({
            "target_id": raw[id_col].astype(str).str.strip(),
            "x": pd.to_numeric(raw[x_col], errors="coerce"),
            "y": pd.to_numeric(raw[y_col], errors="coerce"),
            "kind": kind,
        })

        temp = temp.dropna(subset=["target_id", "x", "y"])
        temp = temp[temp["target_id"] != ""]
        return temp.reset_index(drop=True)

    return clean(sheets[shoot_sheet], "射击"), clean(sheets[photo_sheet], "拍照")


# ============================================================
# 4. 候选执行窗生成
# ============================================================

def quality_score(task_type: str, distance: float, speed: float, accel: float) -> float:
    if task_type == "射击":
        d_min, d_max = SHOOT_D_MIN, SHOOT_D_MAX
        v_max, a_max = SHOOT_V_MAX, SHOOT_A_MAX
    else:
        d_min, d_max = PHOTO_D_MIN, PHOTO_D_MAX
        v_max, a_max = PHOTO_V_MAX, PHOTO_A_MAX

    d_mid = 0.5 * (d_min + d_max)
    d_half = 0.5 * (d_max - d_min)

    q_d = 1.0 - abs(distance - d_mid) / max(d_half, 1e-12)
    q_v = 1.0 - speed / max(v_max, 1e-12)
    q_a = 1.0 - accel / max(a_max, 1e-12)

    return float(max(0.0, 0.45 * q_d + 0.35 * q_v + 0.20 * q_a))


def base_feasible_mask(traj: pd.DataFrame, tx: float, ty: float, task_type: str) -> tuple[np.ndarray, np.ndarray]:
    x = traj["x"].to_numpy(float)
    y = traj["y"].to_numpy(float)
    speed = traj["speed"].to_numpy(float)
    accel = traj["accel"].to_numpy(float)

    dist = np.sqrt((x - tx) ** 2 + (y - ty) ** 2)

    if task_type == "射击":
        mask = (
            (dist >= SHOOT_D_MIN)
            & (dist <= SHOOT_D_MAX)
            & (speed <= SHOOT_V_MAX)
            & (accel <= SHOOT_A_MAX)
        )
    else:
        mask = (
            (dist >= PHOTO_D_MIN)
            & (dist <= PHOTO_D_MAX)
            & (speed <= PHOTO_V_MAX)
            & (accel <= PHOTO_A_MAX)
        )

    return mask, dist


def select_shoot_indices(traj: pd.DataFrame, ready: np.ndarray, dist: np.ndarray) -> list[int]:
    t = traj["time"].to_numpy(float)
    speed = traj["speed"].to_numpy(float)
    accel = traj["accel"].to_numpy(float)
    step_n = max(1, int(round(SHOOT_STEP_S * SCAN_HZ)))

    pool: list[int] = []
    for s, e in continuous_true_blocks(ready):
        idxs = np.arange(s, e + 1, step_n, dtype=int)
        if len(idxs) == 0 or idxs[-1] != e:
            idxs = np.r_[idxs, e]
        pool.extend(idxs.tolist())

    ranked = []
    for i in sorted(set(pool)):
        q = quality_score("射击", dist[i], speed[i], accel[i])
        ranked.append((q, -t[i], i))

    ranked.sort(reverse=True)
    return sorted({i for _, _, i in ranked[:SHOOT_MAX_CAND_PER_TARGET]})


def select_photo_indices(traj: pd.DataFrame, ready: np.ndarray, dist: np.ndarray, tx: float, ty: float) -> list[int]:
    t = traj["time"].to_numpy(float)
    x = traj["x"].to_numpy(float)
    y = traj["y"].to_numpy(float)
    speed = traj["speed"].to_numpy(float)
    accel = traj["accel"].to_numpy(float)
    step_n = max(1, int(round(PHOTO_STEP_S * SCAN_HZ)))

    pool: list[int] = []
    for s, e in continuous_true_blocks(ready):
        idxs = np.arange(s, e + 1, step_n, dtype=int)
        if len(idxs) == 0 or idxs[-1] != e:
            idxs = np.r_[idxs, e]
        pool.extend(idxs.tolist())

    pool = sorted(set(pool))
    if not pool:
        return []

    by_bin: dict[int, list[tuple[float, int]]] = defaultdict(list)
    global_rank = []

    for i in pool:
        ang = bearing_deg(x[i], y[i], tx, ty)
        bin_id = int(math.floor(ang / PHOTO_ANGLE_BIN))
        q = quality_score("拍照", dist[i], speed[i], accel[i])
        by_bin[bin_id].append((q, i))
        global_rank.append((q, -t[i], i))

    selected: list[int] = []

    for arr in by_bin.values():
        arr.sort(reverse=True)
        selected.extend([i for _, i in arr[:3]])

    global_rank.sort(reverse=True)
    selected.extend([i for _, _, i in global_rank[:PHOTO_MAX_CAND_PER_TARGET]])
    selected = sorted(set(selected))

    if len(selected) <= PHOTO_MAX_CAND_PER_TARGET:
        return selected

    ranked = []
    for i in selected:
        ang = bearing_deg(x[i], y[i], tx, ty)
        q = quality_score("拍照", dist[i], speed[i], accel[i])
        ranked.append((q, ang, i))
    ranked.sort(reverse=True)

    kept: list[int] = []
    used_bins = set()
    for _, ang, i in ranked:
        b = int(math.floor(ang / PHOTO_ANGLE_BIN))
        if b not in used_bins:
            kept.append(i)
            used_bins.add(b)
        if len(kept) >= PHOTO_MAX_CAND_PER_TARGET:
            break

    if len(kept) < PHOTO_MAX_CAND_PER_TARGET:
        for _, _, i in ranked:
            if i not in kept:
                kept.append(i)
            if len(kept) >= PHOTO_MAX_CAND_PER_TARGET:
                break

    return sorted(set(kept))


def build_task_windows(traj: pd.DataFrame, shoot_targets: pd.DataFrame, photo_targets: pd.DataFrame) -> tuple[list[TaskWindow], pd.DataFrame]:
    t = traj["time"].to_numpy(float)
    x = traj["x"].to_numpy(float)
    y = traj["y"].to_numpy(float)
    speed = traj["speed"].to_numpy(float)
    accel = traj["accel"].to_numpy(float)

    windows: list[TaskWindow] = []
    diag_rows = []
    uid = 0

    def add_target(row: pd.Series, task_type: str) -> None:
        nonlocal uid

        tid = str(row["target_id"])
        tx = float(row["x"])
        ty = float(row["y"])

        mask, dist = base_feasible_mask(traj, tx, ty, task_type)
        prep = SHOOT_PREP if task_type == "射击" else PHOTO_PREP
        prep_n = max(1, int(math.ceil(prep * SCAN_HZ)))
        ready = all_true_in_previous_window(mask, prep_n)

        if task_type == "射击":
            idxs = select_shoot_indices(traj, ready, dist)
        else:
            idxs = select_photo_indices(traj, ready, dist, tx, ty)

        blocks = continuous_true_blocks(ready)
        total_ready = sum((e - s + 1) / SCAN_HZ for s, e in blocks)

        angle_span = np.nan
        if task_type == "拍照":
            ready_idx = np.where(ready)[0]
            if len(ready_idx):
                angle_span = circular_span_deg(np.array([bearing_deg(x[i], y[i], tx, ty) for i in ready_idx]))

        diag_rows.append({
            "target_id": tid,
            "task_type": task_type,
            "num_ready_intervals": len(blocks),
            "total_ready_duration_s": total_ready,
            "num_candidates": len(idxs),
            "min_distance_m": float(np.nanmin(dist)) if len(dist) else np.nan,
            "photo_ready_angle_span_deg": angle_span,
        })

        for i in idxs:
            ang = bearing_deg(x[i], y[i], tx, ty)
            q = quality_score(task_type, dist[i], speed[i], accel[i])
            windows.append(TaskWindow(
                uid=uid,
                target_id=tid,
                task_type=task_type,
                prep_start=float(t[i] - prep),
                execute_time=float(t[i]),
                occupy_end=float(t[i]),
                robot_x=float(x[i]),
                robot_y=float(y[i]),
                target_x=tx,
                target_y=ty,
                distance=float(dist[i]),
                speed=float(speed[i]),
                accel=float(accel[i]),
                view_angle=float(ang),
                quality=float(q),
            ))
            uid += 1

    for _, row in shoot_targets.iterrows():
        add_target(row, "射击")
    for _, row in photo_targets.iterrows():
        add_target(row, "拍照")

    return windows, pd.DataFrame(diag_rows)


# ============================================================
# 5. 冲突图与整数规划
# ============================================================

def window_conflict(a: TaskWindow, b: TaskWindow) -> bool:
    if a.task_type == "射击" and b.task_type == "射击" and a.target_id == b.target_id:
        return True

    if a.task_type == "拍照" and b.task_type == "拍照" and a.target_id == b.target_id:
        if angle_diff_deg(a.view_angle, b.view_angle) + 1e-9 < PHOTO_ANGLE_MIN:
            return True

    # 准备占用区间不能重叠
    return (b.prep_start < a.occupy_end - 1e-10) and (a.prep_start < b.occupy_end - 1e-10)


def build_conflict_edges(windows: list[TaskWindow]) -> set[tuple[int, int]]:
    n = len(windows)
    edges: set[tuple[int, int]] = set()

    order = sorted(range(n), key=lambda i: windows[i].prep_start)
    for pos, i in enumerate(order):
        wi = windows[i]
        jpos = pos + 1
        while jpos < n:
            j = order[jpos]
            wj = windows[j]
            if wj.prep_start >= wi.occupy_end - 1e-10:
                break
            if window_conflict(wi, wj):
                edges.add(tuple(sorted((i, j))))
            jpos += 1

    photo_by_target = defaultdict(list)
    for i, w in enumerate(windows):
        if w.task_type == "拍照":
            photo_by_target[w.target_id].append(i)

    for idxs in photo_by_target.values():
        for p in range(len(idxs)):
            for q in range(p + 1, len(idxs)):
                i, j = idxs[p], idxs[q]
                if angle_diff_deg(windows[i].view_angle, windows[j].view_angle) + 1e-9 < PHOTO_ANGLE_MIN:
                    edges.add(tuple(sorted((i, j))))

    shoot_by_target = defaultdict(list)
    for i, w in enumerate(windows):
        if w.task_type == "射击":
            shoot_by_target[w.target_id].append(i)
    for idxs in shoot_by_target.values():
        for p in range(len(idxs)):
            for q in range(p + 1, len(idxs)):
                edges.add(tuple(sorted((idxs[p], idxs[q]))))

    return edges


def local_score(w: TaskWindow) -> float:
    return 1000.0 * w.quality + (200.0 if w.task_type == "射击" else 100.0) - 0.0001 * w.execute_time


def can_insert(w: TaskWindow, selected: list[TaskWindow]) -> bool:
    return all(not window_conflict(w, s) for s in selected)


def greedy_fill(selected: list[TaskWindow], all_windows: list[TaskWindow]) -> list[TaskWindow]:
    selected = list(selected)
    used = {w.uid for w in selected}
    pool = [w for w in all_windows if w.uid not in used]
    pool.sort(key=local_score, reverse=True)
    pool = pool[:LOCAL_POOL_SIZE]

    changed = True
    while changed:
        changed = False
        for w in pool:
            if w.uid in used:
                continue
            if can_insert(w, selected):
                selected.append(w)
                used.add(w.uid)
                changed = True
        if changed:
            selected.sort(key=lambda z: z.execute_time)

    return selected


def swap_one_for_two(selected: list[TaskWindow], all_windows: list[TaskWindow]) -> list[TaskWindow]:
    selected_ids = {w.uid for w in selected}
    unselected = [w for w in all_windows if w.uid not in selected_ids]
    unselected.sort(key=local_score, reverse=True)
    unselected = unselected[:LOCAL_POOL_SIZE]

    for rem in sorted(selected, key=lambda w: (w.quality, -w.execute_time)):
        base = [w for w in selected if w.uid != rem.uid]
        addable = []
        for w in unselected:
            if can_insert(w, base):
                addable.append(w)
            if len(addable) >= LOCAL_SWAP_POOL_SIZE:
                break
        if len(addable) < 2:
            continue

        for i in range(len(addable)):
            for j in range(i + 1, len(addable)):
                a, b = addable[i], addable[j]
                if window_conflict(a, b):
                    continue
                trial = base + [a, b]
                trial.sort(key=lambda z: z.execute_time)
                return trial

    return selected


def improve_solution(selected: list[TaskWindow], all_windows: list[TaskWindow]) -> list[TaskWindow]:
    before = len(selected)
    selected = greedy_fill(selected, all_windows)

    improved = True
    while improved:
        old = len(selected)
        selected = swap_one_for_two(selected, all_windows)
        selected = greedy_fill(selected, all_windows)
        improved = len(selected) > old

    selected.sort(key=lambda z: z.execute_time)
    after = len(selected)
    if after > before:
        print(f"局部补插提升：{before} -> {after}")
    else:
        print("局部补插未继续增加任务数。")
    return selected


def solve_task_selection(windows: list[TaskWindow]) -> list[TaskWindow]:
    if not windows:
        raise RuntimeError("未生成任何可执行候选窗。")

    n = len(windows)
    photo_targets = sorted({w.target_id for w in windows if w.task_type == "拍照"})
    y_index = {tid: n + k for k, tid in enumerate(photo_targets)}
    n_var = n + len(photo_targets)

    rows, cols, data = [], [], []
    lb, ub = [], []

    def add_constraint(coeffs: dict[int, float], lo: float, hi: float) -> None:
        r = len(lb)
        for c, v in coeffs.items():
            rows.append(r)
            cols.append(c)
            data.append(v)
        lb.append(lo)
        ub.append(hi)

    # 拍照覆盖变量
    photo_by_target = defaultdict(list)
    for i, w in enumerate(windows):
        if w.task_type == "拍照":
            photo_by_target[w.target_id].append(i)

    for tid, idxs in photo_by_target.items():
        y = y_index[tid]
        for i in idxs:
            add_constraint({i: 1.0, y: -1.0}, -np.inf, 0.0)
        coeffs = {i: -1.0 for i in idxs}
        coeffs[y] = 1.0
        add_constraint(coeffs, -np.inf, 0.0)

    # 冲突约束
    edges = build_conflict_edges(windows)
    for i, j in sorted(edges):
        add_constraint({i: 1.0, j: 1.0}, -np.inf, 1.0)

    print("MILP候选窗数量：", n)
    print("MILP拍照覆盖变量：", len(photo_targets))
    print("冲突边数量：", len(edges))
    print("基础约束数量：", len(lb))

    bounds = Bounds(np.zeros(n_var), np.ones(n_var))
    integrality = np.ones(n_var, dtype=int)

    def run(score: np.ndarray, extra: list[tuple[dict[int, float], float, float]], label: str):
        rr, cc, dd = list(rows), list(cols), list(data)
        ll, uu = list(lb), list(ub)

        for coeffs, lo, hi in extra:
            r = len(ll)
            for c, v in coeffs.items():
                rr.append(r)
                cc.append(c)
                dd.append(v)
            ll.append(lo)
            uu.append(hi)

        A = coo_matrix((dd, (rr, cc)), shape=(len(ll), n_var)).tocsr()
        print(f"开始求解：{label}，约束数={A.shape[0]}")

        res = milp(
            c=-score,
            integrality=integrality,
            bounds=bounds,
            constraints=LinearConstraint(A, np.array(ll), np.array(uu)),
            options={"time_limit": MILP_TIME_LIMIT, "mip_rel_gap": 0.0, "disp": False},
        )

        if not res.success:
            warnings.warn(f"{label} 未完全证明最优：status={res.status}, message={res.message}")
        if res.x is None:
            raise RuntimeError(f"{label} 未返回可行解。")
        return res

    # 第一阶段：最大化总任务数
    score1 = np.zeros(n_var)
    score1[:n] = 1.0
    res1 = run(score1, [], "第一阶段最大任务数")
    max_count = int(round(np.sum(res1.x[:n] > 0.5)))
    print(f"第一阶段任务数：{max_count}")

    # 第二阶段：不降低任务数，优化结构
    t_min = min(w.execute_time for w in windows)
    t_max = max(w.execute_time for w in windows)
    t_span = max(t_max - t_min, 1.0)

    score2 = np.zeros(n_var)
    for i, w in enumerate(windows):
        score2[i] += W_TOTAL
        if w.task_type == "射击":
            score2[i] += W_SHOOT_COUNT
        score2[i] += W_QUALITY * w.quality
        score2[i] -= W_EARLY * ((w.execute_time - t_min) / t_span)

    for y in y_index.values():
        score2[y] += W_PHOTO_COVER

    total_con = ({i: 1.0 for i in range(n)}, float(max_count), np.inf)
    res2 = run(score2, [total_con], "第二阶段同任务数下优化结构")

    chosen_idx = np.where(np.asarray(res2.x[:n]) > 0.5)[0].tolist()
    selected = [windows[i] for i in chosen_idx]
    selected.sort(key=lambda w: w.execute_time)
    print(f"第二阶段任务数：{len(selected)}")

    selected = improve_solution(selected, windows)
    return selected


# ============================================================
# 6. 校验与输出表
# ============================================================

def make_schedule_df(selected: list[TaskWindow]) -> pd.DataFrame:
    rows = []
    for i, w in enumerate(selected, start=1):
        rows.append({
            "序号": i,
            "目标编号": w.target_id,
            "任务类型": w.task_type,
            "开始准备时刻(s)": round(w.prep_start, 3),
            "任务执行时刻(s)": round(w.execute_time, 3),
            "占用结束时刻(s)": round(w.occupy_end, 3),
            "机器人X(m)": round(w.robot_x, 6),
            "机器人Y(m)": round(w.robot_y, 6),
            "目标X(m)": round(w.target_x, 6),
            "目标Y(m)": round(w.target_y, 6),
            "距离(m)": round(w.distance, 6),
            "速度(m/s)": round(w.speed, 6),
            "加速度(m/s²)": round(w.accel, 6),
            "方向角(°)": round(w.view_angle, 6),
            "质量评分": round(w.quality, 6),
        })
    return pd.DataFrame(rows)


def validate_selection(selected: list[TaskWindow]) -> pd.DataFrame:
    errors = []

    for i in range(len(selected)):
        for j in range(i + 1, len(selected)):
            if window_conflict(selected[i], selected[j]):
                errors.append(f"任务冲突：{selected[i].target_id}-{selected[i].task_type} 与 {selected[j].target_id}-{selected[j].task_type}")

    for w in selected:
        if w.task_type == "射击":
            if not (SHOOT_D_MIN <= w.distance <= SHOOT_D_MAX):
                errors.append(f"{w.target_id} 射击距离违规：{w.distance}")
            if w.speed > SHOOT_V_MAX + 1e-9:
                errors.append(f"{w.target_id} 射击速度违规：{w.speed}")
            if w.accel > SHOOT_A_MAX + 1e-9:
                errors.append(f"{w.target_id} 射击加速度违规：{w.accel}")
        else:
            if not (PHOTO_D_MIN <= w.distance <= PHOTO_D_MAX):
                errors.append(f"{w.target_id} 拍照距离违规：{w.distance}")
            if w.speed > PHOTO_V_MAX + 1e-9:
                errors.append(f"{w.target_id} 拍照速度违规：{w.speed}")
            if w.accel > PHOTO_A_MAX + 1e-9:
                errors.append(f"{w.target_id} 拍照加速度违规：{w.accel}")

    photo_by_target = defaultdict(list)
    for w in selected:
        if w.task_type == "拍照":
            photo_by_target[w.target_id].append(w)

    for tid, arr in photo_by_target.items():
        for i in range(len(arr)):
            for j in range(i + 1, len(arr)):
                d = angle_diff_deg(arr[i].view_angle, arr[j].view_angle)
                if d + 1e-9 < PHOTO_ANGLE_MIN:
                    errors.append(f"{tid} 拍照角度差违规：{d:.3f}°")

    return pd.DataFrame({"error": errors})


def make_photo_check_df(selected: list[TaskWindow], photo_targets: pd.DataFrame) -> pd.DataFrame:
    by_target = defaultdict(list)
    for w in selected:
        if w.task_type == "拍照":
            by_target[w.target_id].append(w.view_angle)

    rows = []
    for _, row in photo_targets.iterrows():
        tid = str(row["target_id"])
        angles = sorted(by_target.get(tid, []))
        min_diff = np.nan
        if len(angles) >= 2:
            diffs = [angle_diff_deg(angles[i], angles[j]) for i in range(len(angles)) for j in range(i + 1, len(angles))]
            min_diff = float(min(diffs))
        rows.append({
            "target_id": tid,
            "num_photos": len(angles),
            "angles_deg": ", ".join(f"{a:.2f}" for a in angles),
            "min_pairwise_angle_diff_deg": min_diff,
            "required_min_angle_diff_deg": PHOTO_ANGLE_MIN,
            "pass": True if len(angles) <= 1 else (min_diff + 1e-9 >= PHOTO_ANGLE_MIN),
        })
    return pd.DataFrame(rows)


def make_summary_df(schedule: pd.DataFrame, shoot_targets: pd.DataFrame, photo_targets: pd.DataFrame) -> pd.DataFrame:
    shoot_times = int((schedule["任务类型"] == "射击").sum()) if not schedule.empty else 0
    photo_times = int((schedule["任务类型"] == "拍照").sum()) if not schedule.empty else 0
    shoot_done = schedule[schedule["任务类型"] == "射击"]["目标编号"].nunique() if not schedule.empty else 0
    photo_done = schedule[schedule["任务类型"] == "拍照"]["目标编号"].nunique() if not schedule.empty else 0

    return pd.DataFrame([
        ["射击目标总数", len(shoot_targets)],
        ["拍照目标总数", len(photo_targets)],
        ["完成射击目标数", shoot_done],
        ["安排射击次数", shoot_times],
        ["射击期望命中数", round(SHOOT_HIT_PROB * shoot_times, 3)],
        ["覆盖拍照目标数", photo_done],
        ["安排拍照总次数", photo_times],
        ["总任务行数", len(schedule)],
        ["拍照最小角度差要求(°)", PHOTO_ANGLE_MIN],
    ], columns=["指标", "值"])


# ============================================================
# 7. 写 result.xlsx，保护红字
# ============================================================

def is_red_font(cell) -> bool:
    color = cell.font.color
    if color is None:
        return False
    if color.type == "rgb" and color.rgb:
        rgb = color.rgb.upper()
        return rgb.endswith("FF0000") or rgb.endswith("C00000") or rgb.endswith("E60000")
    if color.type == "indexed" and color.indexed in {3, 10}:
        return True
    return False


def safe_write(cell, value) -> None:
    if is_red_font(cell):
        raise RuntimeError(f"拒绝写入红字单元格：{cell.coordinate}")
    cell.value = value


def safe_clear(cell) -> None:
    if not is_red_font(cell):
        cell.value = None


def find_result_template() -> Path | None:
    for p in RESULT_TEMPLATE_CANDIDATES:
        if p.exists():
            return p
    return None


def find_header_row(ws) -> int | None:
    for r in range(1, min(ws.max_row, 100) + 1):
        vals = [norm_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        row_text = "".join(vals)
        if ("目标" in row_text or "编号" in row_text) and ("时刻" in row_text or "时间" in row_text or "任务" in row_text):
            if sum(1 for v in vals if v) >= 3:
                return r
    return None


def header_map(ws, header_row: int) -> dict[str, int]:
    ans = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is not None:
            ans[norm_text(v)] = c
    return ans


def match_header(headers: dict[str, int], aliases: list[str]) -> int | None:
    alias_norm = [norm_text(a) for a in aliases]
    for a in alias_norm:
        if a in headers:
            return headers[a]
    for h, c in headers.items():
        for a in alias_norm:
            if a in h or h in a:
                return c
    return None


def write_result_xlsx(schedule: pd.DataFrame, summary: pd.DataFrame, photo_check: pd.DataFrame) -> Path:
    out_path = OUT_DIR / "result.xlsx"

    template = find_result_template()
    if template:
        shutil.copy2(template, out_path)
        wb = load_workbook(out_path)
    else:
        wb = Workbook()

    target_ws = None
    header_row = None
    for ws in wb.worksheets:
        hr = find_header_row(ws)
        if hr is not None:
            target_ws = ws
            header_row = hr
            break

    if target_ws is None:
        target_ws = wb.create_sheet("任务调度结果")
        header_row = 1
        for c, name in enumerate(schedule.columns, start=1):
            safe_write(target_ws.cell(header_row, c), name)

    headers = header_map(target_ws, header_row)
    col_alias = {
        "序号": ["序号", "序", "编号"],
        "目标编号": ["目标编号", "目标", "目标id", "target_id"],
        "任务类型": ["任务类型", "任务", "类型"],
        "开始准备时刻(s)": ["开始准备时刻(s)", "准备开始时刻", "开始时间"],
        "任务执行时刻(s)": ["任务执行时刻(s)", "执行时刻", "拍摄时刻", "射击时刻", "执行时间"],
        "占用结束时刻(s)": ["结束时刻", "占用结束时刻", "结束时间"],
        "距离(m)": ["距离(m)", "距离"],
        "速度(m/s)": ["速度(m/s)", "速度"],
        "加速度(m/s²)": ["加速度(m/s²)", "加速度(m/s2)", "加速度"],
        "方向角(°)": ["方向角(°)", "方向角", "拍摄方向角", "角度"],
    }

    write_cols = {}
    for df_col, aliases in col_alias.items():
        if df_col in schedule.columns:
            col = match_header(headers, aliases)
            if col is not None:
                write_cols[df_col] = col

    if len(write_cols) < 3:
        target_ws = wb.create_sheet("任务调度结果_完整")
        header_row = 1
        for c, name in enumerate(schedule.columns, start=1):
            safe_write(target_ws.cell(header_row, c), name)
        write_cols = {name: i + 1 for i, name in enumerate(schedule.columns)}

    start_row = header_row + 1
    for r in range(start_row, start_row + max(300, len(schedule) + 50)):
        for c in write_cols.values():
            safe_clear(target_ws.cell(r, c))

    for ridx, (_, row) in enumerate(schedule.iterrows(), start=start_row):
        for df_col, col in write_cols.items():
            safe_write(target_ws.cell(ridx, col), row[df_col])

    for name in ["第四问汇总", "拍照角度检查"]:
        if name in wb.sheetnames:
            del wb[name]

    ws_sum = wb.create_sheet("第四问汇总")
    for c, name in enumerate(summary.columns, start=1):
        ws_sum.cell(1, c).value = name
    for r, (_, row) in enumerate(summary.iterrows(), start=2):
        for c, value in enumerate(row, start=1):
            ws_sum.cell(r, c).value = value

    ws_ang = wb.create_sheet("拍照角度检查")
    for c, name in enumerate(photo_check.columns, start=1):
        ws_ang.cell(1, c).value = name
    for r, (_, row) in enumerate(photo_check.iterrows(), start=2):
        for c, value in enumerate(row, start=1):
            ws_ang.cell(r, c).value = value

    wb.save(out_path)
    return out_path


# ============================================================
# 8. 作图
# ============================================================

def setup_chinese_font() -> None:
    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei", "Noto Sans CJK SC", "Arial Unicode MS", "DejaVu Sans"]


def plot_results(traj: pd.DataFrame, shoot_targets: pd.DataFrame, photo_targets: pd.DataFrame, selected: list[TaskWindow], photo_check: pd.DataFrame) -> None:
    setup_chinese_font()
    schedule = make_schedule_df(selected)

    plt.figure(figsize=(10, 9))
    sc = plt.scatter(traj["x"], traj["y"], c=traj["speed"], s=4, alpha=0.65)
    plt.colorbar(sc, label="速度/(m/s)")
    plt.scatter(shoot_targets["x"], shoot_targets["y"], marker="x", s=75, label="射击目标")
    plt.scatter(photo_targets["x"], photo_targets["y"], marker="o", s=65, facecolors="none", label="拍照目标")

    if not schedule.empty:
        ssel = schedule[schedule["任务类型"] == "射击"]
        psel = schedule[schedule["任务类型"] == "拍照"]
        if not ssel.empty:
            plt.scatter(ssel["目标X(m)"], ssel["目标Y(m)"], marker="*", s=130, label="已安排射击")
        if not psel.empty:
            plt.scatter(psel["目标X(m)"], psel["目标Y(m)"], marker=".", s=80, label="已安排拍照")

    plt.axis("equal")
    plt.xlabel("X坐标/m")
    plt.ylabel("Y坐标/m")
    plt.title("第四问候选执行窗调度结果")
    plt.grid(True, alpha=0.3)
    plt.legend()
    plt.tight_layout()
    plt.savefig(OUT_DIR / "q4_trajectory_targets_second.png", dpi=300)
    plt.close()

    plt.figure(figsize=(11, 5))
    tmp = photo_check.copy()
    tmp["num_photos"] = tmp["num_photos"].astype(int)
    plt.bar(tmp["target_id"], tmp["num_photos"])
    plt.xlabel("拍照目标编号")
    plt.ylabel("拍照次数")
    plt.title(f"各拍照目标拍照次数（方向角差 ≥ {PHOTO_ANGLE_MIN:g}°）")
    plt.grid(True, axis="y", alpha=0.3)
    plt.tight_layout()
    plt.savefig(OUT_DIR / "q4_photo_counts_second.png", dpi=300)
    plt.close()


# ============================================================
# 9. 主程序
# ============================================================

def main() -> None:
    ensure_output_dir()

    print("=" * 72)
    print("第四问第二版：候选执行窗—冲突图—整数规划调度")
    print("=" * 72)

    traj_path = find_q3_trajectory()
    target_path = find_first_existing(TARGET_FILE_CANDIDATES, "附件4目标文件")

    print(f"第三问轨迹文件：{traj_path}")
    print(f"附件4目标文件：{target_path}")
    print(f"拍照最小方向角差：{PHOTO_ANGLE_MIN:g}°")

    traj_10hz = read_trajectory(traj_path)
    traj_dense = densify_trajectory(traj_10hz)

    print(f"第三问轨迹点数：{len(traj_10hz)}")
    print(f"高密度扫描点数：{len(traj_dense)}")
    print(f"扫描时间区间：{traj_dense['time'].iloc[0]:.3f} — {traj_dense['time'].iloc[-1]:.3f} s")
    print(f"速度范围：{traj_dense['speed'].min():.3f} — {traj_dense['speed'].max():.3f} m/s")
    print(f"加速度范围：{traj_dense['accel'].min():.3f} — {traj_dense['accel'].max():.3f} m/s²")

    shoot_targets, photo_targets = read_targets(target_path)
    print(f"射击目标数：{len(shoot_targets)}")
    print(f"拍照目标数：{len(photo_targets)}")

    windows, diagnosis = build_task_windows(traj_dense, shoot_targets, photo_targets)
    print(f"候选执行窗数量：{len(windows)}")

    selected = solve_task_selection(windows)
    schedule = make_schedule_df(selected)

    validation = validate_selection(selected)
    if not validation.empty:
        err_path = OUT_DIR / "q4_validation_errors.csv"
        validation.to_csv(err_path, index=False, encoding="utf-8-sig")
        raise RuntimeError(f"调度结果存在违规，已保存：{err_path}")

    photo_check = make_photo_check_df(selected, photo_targets)
    summary = make_summary_df(schedule, shoot_targets, photo_targets)

    selected_count = (
        schedule.groupby(["目标编号", "任务类型"]).size().reset_index(name="selected_count")
        if not schedule.empty else
        pd.DataFrame(columns=["目标编号", "任务类型", "selected_count"])
    )
    diagnosis = diagnosis.merge(selected_count, how="left", left_on=["target_id", "task_type"], right_on=["目标编号", "任务类型"])
    diagnosis["selected_count"] = diagnosis["selected_count"].fillna(0).astype(int)
    diagnosis = diagnosis.drop(columns=[c for c in ["目标编号", "任务类型"] if c in diagnosis.columns])

    schedule_path = OUT_DIR / "q4_task_schedule_second.csv"
    diagnosis_path = OUT_DIR / "q4_target_diagnosis_second.csv"
    angle_path = OUT_DIR / "q4_photo_angle_check_second.csv"
    summary_path = OUT_DIR / "q4_summary_second.csv"
    dense_path = OUT_DIR / "q4_dense_trajectory_second.csv"

    schedule.to_csv(schedule_path, index=False, encoding="utf-8-sig")
    diagnosis.to_csv(diagnosis_path, index=False, encoding="utf-8-sig")
    photo_check.to_csv(angle_path, index=False, encoding="utf-8-sig")
    summary.to_csv(summary_path, index=False, encoding="utf-8-sig")
    traj_dense.to_csv(dense_path, index=False, encoding="utf-8-sig")

    with pd.ExcelWriter(OUT_DIR / "q4_debug_tables_second.xlsx", engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="汇总", index=False)
        schedule.to_excel(writer, sheet_name="任务明细", index=False)
        diagnosis.to_excel(writer, sheet_name="目标诊断", index=False)
        photo_check.to_excel(writer, sheet_name="拍照角度检查", index=False)

    result_path = write_result_xlsx(schedule, summary, photo_check)
    plot_results(traj_dense, shoot_targets, photo_targets, selected, photo_check)

    shoot_done = schedule[schedule["任务类型"] == "射击"]["目标编号"].nunique() if not schedule.empty else 0
    shoot_times = int((schedule["任务类型"] == "射击").sum()) if not schedule.empty else 0
    photo_done = schedule[schedule["任务类型"] == "拍照"]["目标编号"].nunique() if not schedule.empty else 0
    photo_times = int((schedule["任务类型"] == "拍照").sum()) if not schedule.empty else 0

    print("\n" + "=" * 72)
    print("第四问求解完成")
    print("=" * 72)
    print(f"射击目标完成数：{shoot_done} / {len(shoot_targets)}")
    print(f"安排射击次数：{shoot_times}")
    print(f"射击期望命中数：{SHOOT_HIT_PROB * shoot_times:.3f}")
    print(f"拍照目标覆盖数：{photo_done} / {len(photo_targets)}")
    print(f"安排拍照总次数：{photo_times}")
    print(f"总任务行数：{len(schedule)}")
    print(f"拍照角度约束全部通过：{bool(photo_check['pass'].all())}")

    print("\n输出文件：")
    print(f"提交表：{result_path}")
    print(f"任务明细：{schedule_path}")
    print(f"目标诊断：{diagnosis_path}")
    print(f"拍照角度检查：{angle_path}")
    print(f"汇总表：{summary_path}")
    print(f"图件目录：{OUT_DIR}")

    print("\n说明：")
    print("1. 该版本使用候选执行窗—冲突图—整数规划框架。")
    print("2. 准备时间段内约束采用 all_true 严格判断，没有使用90%放宽。")
    print("3. 同一拍照目标可多次出现，但任意两次方向角差必须不少于60°。")
    print("4. 程序保护 result.xlsx 红字单元格。")


if __name__ == "__main__":
    main()
