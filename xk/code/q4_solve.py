"""Q4: 基于附件3 轨迹 (Q3 KF/RTS 输出) 的多目标任务调度优化.

题面约束 (严格按题目附录):
  射击: d ∈ [5, 30] m, v ≤ 2 m/s, |a| ≤ 1.5 m/s², 校准 1.5s 持续满足, 命中率 85%
  拍照: d ∈ [10, 40] m, v ≤ 1.5 m/s, |a| ≤ 1.5 m/s², 对准 0.5s 持续满足
        同一目标多角度: 方向差 ≥ 60°

策略:
  1) 加载 Q3_trajectory_10Hz_kalman.xlsx (KF/RTS 输出, 含速度 + 后验方差)
  2) 求加速度 (KF 速度的中心差分)
  3) 对每个目标 (S01-S18, P01-P18) 找候选可执行段:
     - 滑动窗判定: 起始 j, 窗 [j, j+W) 全满足约束
     - W = 15 (射击 1.5s) / 5 (拍照 0.5s)
  4) 段内取距离最小点为执行时刻 (对射击)
     对拍照, 同目标分多角度: 按视角聚类 60° bins, 每 bin 取段距离最小点
  5) 调度: 加权区间调度 DP — 每任务权重 1, 最大化总任务数
     约束: 准备-执行时段 [t_prep, t_exec] 不重叠
  6) 输出: result.xlsx (按模板格式, 含序号/目标/任务/准备时刻/执行时刻)

输出:
  xk/output/Q4_summary.json
  xk/output/Q4_全部候选.xlsx
  xk/output/Q4_最终方案.xlsx
  xk/output/Q4_result_filled.xlsx (按 result_template 格式填充)
  xk/figures/Q4_overview.png
  xk/figures/Q4_gantt.png
  xk/figures/Q4_kinematic_profile.png
"""
from __future__ import annotations

import bisect
import json
import shutil
import sys
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT.parent / "A"))
try:
    from plot_style import setup_plot_style  # noqa: E402
    setup_plot_style()
except Exception:
    from matplotlib import rcParams
    rcParams['font.sans-serif'] = ['Noto Sans CJK SC', 'SimHei']
    rcParams['axes.unicode_minus'] = False

DATA = ROOT / "data"
FIGS = ROOT / "figures"
OUT = ROOT / "output"


# =====================================================================
# 题面约束常数 (严格按 docs/题目.md 附录)
# =====================================================================
# 射击
D_SHOOT_MIN, D_SHOOT_MAX = 5.0, 30.0   # m
V_SHOOT_MAX = 2.0                       # m/s
A_SHOOT_MAX = 1.5                       # m/s²
T_AIM = 1.5                             # s 校准
HIT_RATE = 0.85

# 拍照
D_PHOTO_MIN, D_PHOTO_MAX = 10.0, 40.0  # m
V_PHOTO_MAX = 1.5                       # m/s
A_PHOTO_MAX = 1.5                       # m/s²
T_FOCUS = 0.5                           # s 对准
ANG_DIFF_DEG = 60.0                     # ° 同目标多角度差

# =====================================================================
# 三审 P0 修正: 鲁棒约束 (chance constraint) + 过渡时间
# =====================================================================
# 鲁棒约束 z 值: 1.645 = 90% 单侧, 1.96 = 95%; 三方建议 1.645 (实务推荐)
ROBUST_Z = 1.645
# 速度估计相对不确定度 (Claude 估计 5%)
SIGMA_V_REL = 0.05
# 任务间过渡时间 (武器/相机切换), 三方一致建议 0.1 s
EPSILON_TRANSITION = 0.1


# =====================================================================
# 1. 加载 Q3 KF/RTS 轨迹 + 附件4 目标
# =====================================================================
def load_kinematic():
    """加载 Q3 平滑轨迹 + 求加速度 + KF 后验 σ."""
    df = pd.read_excel(OUT / "Q3_trajectory_10Hz_kalman.xlsx")
    t = df["time_s"].values
    x = df["X_m"].values
    y = df["Y_m"].values
    vx = df["Vx_m_s"].values
    vy = df["Vy_m_s"].values
    var_x = df["var_X"].values
    var_y = df["var_Y"].values
    sigma_x = np.sqrt(var_x)
    sigma_y = np.sqrt(var_y)

    dt = float(np.median(np.diff(t)))

    # 中心差分求加速度
    ax = np.zeros_like(vx); ay = np.zeros_like(vy)
    ax[1:-1] = (vx[2:] - vx[:-2]) / (2 * dt)
    ay[1:-1] = (vy[2:] - vy[:-2]) / (2 * dt)
    ax[0], ax[-1] = ax[1], ax[-2]
    ay[0], ay[-1] = ay[1], ay[-2]

    speed = np.hypot(vx, vy)
    accel = np.hypot(ax, ay)
    # 速度不确定度: σ_v ≈ 5% × |v| (Claude 估计) + KF 速度方差应附加, 此处保守用比例
    sigma_v = SIGMA_V_REL * np.maximum(speed, 0.1)  # 下限 0.1 m/s 避免 σ_v→0

    return dict(t=t, x=x, y=y, vx=vx, vy=vy, ax=ax, ay=ay,
                speed=speed, accel=accel,
                sigma_x=sigma_x, sigma_y=sigma_y, sigma_v=sigma_v,
                dt=dt)


def load_targets():
    """加载 18 射击 + 18 拍照目标."""
    xl = pd.ExcelFile(DATA / "附件4.xlsx")
    df_s = pd.read_excel(xl, "射击目标")
    df_p = pd.read_excel(xl, "拍照目标")
    S = df_s[["X坐标(m)", "Y坐标(m)"]].values
    S_id = df_s["编号"].tolist()
    P = df_p[["X坐标(m)", "Y坐标(m)"]].values
    P_id = df_p["编号"].tolist()
    return S, S_id, P, P_id


# =====================================================================
# 2. 滑动窗找可行段
# =====================================================================
def find_segments(ok_w):
    """从布尔数组找连续 True 段, 返回 [(i_start, i_end), ...] 闭区间."""
    segs = []
    in_seg = False
    s0 = 0
    for j in range(len(ok_w)):
        if ok_w[j] and not in_seg:
            in_seg = True
            s0 = j
        elif not ok_w[j] and in_seg:
            in_seg = False
            segs.append((s0, j - 1))
    if in_seg:
        segs.append((s0, len(ok_w) - 1))
    return segs


def _sigma_d(kin, target):
    """距离不确定度的传播: σ_d² = (∂d/∂x)² σ_x² + (∂d/∂y)² σ_y².
    线性化: ∂d/∂x = (x - x_g)/d, 同理 ∂d/∂y = (y - y_g)/d.
    """
    dx = kin["x"] - target[0]
    dy = kin["y"] - target[1]
    d = np.hypot(dx, dy)
    d_safe = np.maximum(d, 1e-6)
    sigma_d_sq = (dx / d_safe) ** 2 * kin["sigma_x"] ** 2 \
               + (dy / d_safe) ** 2 * kin["sigma_y"] ** 2
    return np.sqrt(sigma_d_sq)


def shoot_candidates(S, S_id, kin, W_aim, robust: bool = False):
    """对每个射击目标找候选执行时刻段, 取段内距离最小点为执行时刻.

    robust=True 时启用 chance constraint:
        d - z·σ_d ≥ d_min (左下界)
        d + z·σ_d ≤ d_max (右上界)
        v + z·σ_v ≤ v_max
    其中 z = ROBUST_Z (1.645, 90% 单侧).
    """
    cands = []
    z = ROBUST_Z if robust else 0.0
    for i, s in enumerate(S):
        d = np.hypot(kin["x"] - s[0], kin["y"] - s[1])
        sigma_d = _sigma_d(kin, s) if robust else np.zeros_like(d)
        d_lo = d - z * sigma_d  # 用于 d ≥ d_min 检查的保守下界
        d_hi = d + z * sigma_d  # 用于 d ≤ d_max 检查的保守上界
        v_hi = kin["speed"] + z * kin["sigma_v"]
        a_hi = kin["accel"]    # σ_a 估计噪声大, 不参与鲁棒
        ok_pt = (d_lo >= D_SHOOT_MIN) & (d_hi <= D_SHOOT_MAX) \
                & (v_hi <= V_SHOOT_MAX) \
                & (a_hi <= A_SHOOT_MAX)
        # 滑窗: j 起 [j, j+W_aim) 全满足
        N = len(d)
        ok_w = np.zeros(N, dtype=bool)
        for j in range(N - W_aim + 1):
            if np.all(ok_pt[j:j + W_aim]):
                ok_w[j] = True
        segs = find_segments(ok_w)
        for (j0, j1) in segs:
            # 执行时刻索引 = 准备开始 + W_aim - 1
            exec_idxs = np.arange(j0 + W_aim - 1, min(j1 + W_aim, N))
            if len(exec_idxs) == 0:
                continue
            d_in = d[exec_idxs]
            kmin = exec_idxs[int(np.argmin(d_in))]
            cands.append({
                "target": S_id[i], "type": "射击",
                "t_exec": float(kin["t"][kmin]),
                "t_prep": float(kin["t"][kmin] - T_AIM),
                "d": float(d[kmin]),
                "v": float(kin["speed"][kmin]),
                "a": float(kin["accel"][kmin]),
                "idx_exec": int(kmin),
                "view_angle": float('nan'),  # 射击无视角
            })
    return cands


def photo_candidates(P, P_id, kin, W_aim, ang_diff_rad, robust: bool = False):
    """对每个拍照目标找多个候选 (≥60° 视角差).

    robust=True 同 shoot_candidates 启用 chance constraint.
    """
    cands = []
    z = ROBUST_Z if robust else 0.0
    for i, p in enumerate(P):
        d = np.hypot(kin["x"] - p[0], kin["y"] - p[1])
        sigma_d = _sigma_d(kin, p) if robust else np.zeros_like(d)
        d_lo = d - z * sigma_d
        d_hi = d + z * sigma_d
        v_hi = kin["speed"] + z * kin["sigma_v"]
        a_hi = kin["accel"]
        ok_pt = (d_lo >= D_PHOTO_MIN) & (d_hi <= D_PHOTO_MAX) \
                & (v_hi <= V_PHOTO_MAX) \
                & (a_hi <= A_PHOTO_MAX)
        N = len(d)
        ok_w = np.zeros(N, dtype=bool)
        for j in range(N - W_aim + 1):
            if np.all(ok_pt[j:j + W_aim]):
                ok_w[j] = True
        segs = find_segments(ok_w)
        # 每段产生一个 "段距离最小" 候选
        target_cands = []
        for (j0, j1) in segs:
            exec_idxs = np.arange(j0 + W_aim - 1, min(j1 + W_aim, N))
            if len(exec_idxs) == 0:
                continue
            d_in = d[exec_idxs]
            kmin = exec_idxs[int(np.argmin(d_in))]
            view_ang = float(np.arctan2(kin["y"][kmin] - p[1], kin["x"][kmin] - p[0]))
            target_cands.append({
                "target": P_id[i], "type": "拍照",
                "t_exec": float(kin["t"][kmin]),
                "t_prep": float(kin["t"][kmin] - T_FOCUS),
                "d": float(d[kmin]),
                "v": float(kin["speed"][kmin]),
                "a": float(kin["accel"][kmin]),
                "idx_exec": int(kmin),
                "view_angle": view_ang,
            })
        # 同目标按时间排序后, 角度差 ≥ ANG_DIFF 才保留
        target_cands.sort(key=lambda c: c["t_exec"])
        chosen = []
        for c in target_cands:
            keep = True
            for ch in chosen:
                d_ang = abs(np.angle(np.exp(1j * (c["view_angle"] - ch["view_angle"]))))
                if d_ang < ang_diff_rad:
                    keep = False
                    break
            if keep:
                chosen.append(c)
        cands.extend(chosen)
    return cands


# =====================================================================
# 3. 加权区间调度 DP (最大化任务数)
# =====================================================================
def weighted_interval_scheduling_with_uniqueness(all_cand, epsilon: float = 0.0):
    """加权区间调度 + 唯一性约束:
       - 射击: 每个目标至多入选 1 次 (完成数 = 不同目标数, 而非射击次数)
       - 拍照: 同目标允许多次 (前提是视角差 ≥ 60°, 已在候选生成阶段去重过)
       - epsilon: 任务间过渡时间 (s), 即 t_exec_i + ε ≤ t_prep_j
    使用 ILP (scipy.milp) 求最优; 退化时用贪心.
    """
    if not all_cand:
        return []
    n = len(all_cand)
    # 排序: 按 t_exec
    cand = sorted(all_cand, key=lambda c: c["t_exec"])
    for i, c in enumerate(cand):
        c["_idx"] = i

    # 用 scipy.milp 求解 ILP (最优)
    try:
        from scipy.optimize import milp, LinearConstraint, Bounds
        from scipy.sparse import lil_matrix
        # 目标: 最大化 ∑ x_i ⇔ 最小化 -∑ x_i
        c_obj = -np.ones(n)
        # 约束矩阵: 收集所有 ≤ 1 的不等式行
        rows = []
        # 时段冲突: t_exec_i + ε > t_prep_j 时互斥 (考虑过渡时间)
        for i in range(n):
            for j in range(i + 1, n):
                if cand[i]["t_exec"] + epsilon > cand[j]["t_prep"]:
                    row = np.zeros(n)
                    row[i] = 1; row[j] = 1
                    rows.append(row)
        # 射击唯一: 同一目标 ≤ 1
        shoot_by_target = {}
        for i, cc in enumerate(cand):
            if cc["type"] == "射击":
                shoot_by_target.setdefault(cc["target"], []).append(i)
        for tgt, idxs in shoot_by_target.items():
            if len(idxs) > 1:
                row = np.zeros(n)
                for i in idxs:
                    row[i] = 1
                rows.append(row)
        if rows:
            A = np.vstack(rows)
            b_ub = np.ones(len(rows))
            constraints = LinearConstraint(A, ub=b_ub)
        else:
            constraints = []
        result = milp(
            c_obj, constraints=constraints,
            integrality=np.ones(n, dtype=int),
            bounds=Bounds(lb=0, ub=1),
        )
        if result.success:
            x_sol = result.x
            sel_idx = [i for i in range(n) if x_sol[i] > 0.5]
            return [cand[i] for i in sel_idx]
        else:
            print(f"  [警告] ILP 求解失败 ({result.message}), 回退到贪心")
    except (ImportError, AttributeError) as e:
        print(f"  [警告] scipy.milp 不可用: {e}, 回退到贪心")
    # 退化贪心 (理论不会执行, ILP 必走) — 仅作防御
    if False:
        # 退化: 贪心 — 先按射击目标去重 (留最佳: d 最小), 再做 DP
        shoot_best = {}
        for c in cand:
            if c["type"] == "射击":
                if c["target"] not in shoot_best or c["d"] < shoot_best[c["target"]]["d"]:
                    shoot_best[c["target"]] = c
        cand2 = [c for c in cand if c["type"] != "射击"] + list(shoot_best.values())
        cand2.sort(key=lambda c: c["t_exec"])
        n2 = len(cand2)
        end_times = [c["t_exec"] for c in cand2]
        p = [-1] * n2
        for i, c in enumerate(cand2):
            j = bisect.bisect_right(end_times, c["t_prep"]) - 1
            p[i] = j
        dp = [0] * (n2 + 1)
        take = [False] * n2
        for i in range(n2):
            inc = 1 + dp[p[i] + 1]
            if inc > dp[i]:
                dp[i + 1] = inc
                take[i] = True
            else:
                dp[i + 1] = dp[i]
        sel_idx = []
        i = n2 - 1
        while i >= 0:
            if take[i]:
                sel_idx.append(i); i = p[i]
            else:
                i -= 1
        sel_idx.sort()
        return [cand2[i] for i in sel_idx]


# =====================================================================
# 4. 输出
# =====================================================================
def write_result_xlsx(selected):
    """按 result_template 格式填充 result_filled.xlsx."""
    template = DATA / "result_template.xlsx"
    out = OUT / "Q4_result_filled.xlsx"
    shutil.copy(template, out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Sheet1"]
    # 清除 A2:E... 旧数据 (仅左侧表)
    max_row = max(ws.max_row, len(selected) + 1)
    for r in range(2, max_row + 2):
        for col_letter in ('A', 'B', 'C', 'D', 'E'):
            ws[f"{col_letter}{r}"] = None
    # 写入
    for i, c in enumerate(selected):
        r = i + 2
        ws.cell(r, 1).value = i + 1
        ws.cell(r, 2).value = c["target"]
        ws.cell(r, 3).value = c["type"]
        ws.cell(r, 4).value = round(c["t_prep"], 2)
        ws.cell(r, 5).value = round(c["t_exec"], 2)
    wb.save(out)
    return out


def plot_overview(kin, S, S_id, P, P_id, all_cand, selected):
    fig, axes = plt.subplots(1, 2, figsize=(13, 6.5))

    # 左: XY 轨迹 + 目标点
    ax = axes[0]
    ax.plot(kin["x"], kin["y"], 'b-', lw=0.6, alpha=0.6, label='机器人轨迹')
    ax.scatter(S[:, 0], S[:, 1], marker='x', s=80, c='red',
               linewidth=1.5, label=f'射击目标 ({len(S)})')
    ax.scatter(P[:, 0], P[:, 1], marker='o', s=80, c='green',
               edgecolor='black', linewidth=1.0, label=f'拍照目标 ({len(P)})')
    # 标注完成的目标
    sel_targets = {c["target"] for c in selected}
    for i, sid in enumerate(S_id):
        if sid in sel_targets:
            ax.annotate(sid, (S[i, 0], S[i, 1]),
                        xytext=(5, 5), textcoords='offset points',
                        fontsize=7, color='red', fontweight='bold')
    for i, pid in enumerate(P_id):
        if pid in sel_targets:
            ax.annotate(pid, (P[i, 0], P[i, 1]),
                        xytext=(5, 5), textcoords='offset points',
                        fontsize=7, color='green', fontweight='bold')
    ax.set_xlabel('X (m)'); ax.set_ylabel('Y (m)')
    ax.set_title(f'附件3 轨迹 + 任务目标 (完成 {len(selected)})')
    ax.legend(loc='upper right', fontsize=9)
    ax.grid(alpha=0.3)
    ax.set_aspect('equal')

    # 右: 速度 / 加速度时序 + 选中任务的执行时刻
    ax = axes[1]
    ax2 = ax.twinx()
    l1 = ax.plot(kin["t"], kin["speed"], 'b-', lw=0.6, alpha=0.7, label='|v| (m/s)')
    l2 = ax2.plot(kin["t"], kin["accel"], 'r-', lw=0.4, alpha=0.5, label='|a| (m/s²)')
    ax.axhline(V_SHOOT_MAX, color='orange', ls=':', lw=1, alpha=0.6,
               label=f'射击 v 上限 ({V_SHOOT_MAX})')
    ax.axhline(V_PHOTO_MAX, color='green', ls=':', lw=1, alpha=0.6,
               label=f'拍照 v 上限 ({V_PHOTO_MAX})')
    for c in selected:
        color = 'red' if c["type"] == '射击' else 'green'
        ax.axvline(c["t_exec"], color=color, lw=0.5, alpha=0.6)
    ax.set_xlabel('时间 (s)')
    ax.set_ylabel('|v| (m/s)', color='b')
    ax2.set_ylabel('|a| (m/s²)', color='r')
    ax.set_title('运动学 + 选中任务执行时刻')
    ax.legend(loc='upper right', fontsize=8)
    ax.grid(alpha=0.3)

    plt.tight_layout()
    plt.savefig(FIGS / 'Q4_overview.png', dpi=200)
    plt.close()


def plot_gantt(selected):
    fig, ax = plt.subplots(figsize=(13, max(6, len(selected) * 0.3)))
    selected_sorted = sorted(selected, key=lambda c: c["t_prep"])
    for i, c in enumerate(selected_sorted):
        y = len(selected_sorted) - i
        color = '#d62728' if c["type"] == '射击' else '#2ca02c'
        # 准备段
        ax.barh(y, c["t_exec"] - c["t_prep"], left=c["t_prep"],
                color=color, alpha=0.55, edgecolor='black', linewidth=0.5)
        # 执行时刻打点
        ax.plot(c["t_exec"], y, 'k|', ms=10)
        ax.text(c["t_prep"] - 1, y, c["target"], ha='right', va='center',
                fontsize=7)
    ax.set_xlabel('时间 (s)')
    ax.set_ylabel('任务序号 (按时间)')
    ax.set_title(f'任务调度甘特图 ({len(selected_sorted)} 任务)')
    ax.grid(axis='x', alpha=0.3)
    plt.tight_layout()
    plt.savefig(FIGS / 'Q4_gantt.png', dpi=200)
    plt.close()


def plot_kinematic_profile(kin, selected):
    """每个选中任务的 d, v, a 时序剖面 (校准期间)."""
    n_sel = len(selected)
    if n_sel == 0:
        return
    n_cols = 4
    n_rows = (n_sel + n_cols - 1) // n_cols
    fig, axes = plt.subplots(n_rows, n_cols, figsize=(13, 2.4 * n_rows),
                             squeeze=False)
    for i, c in enumerate(selected):
        ax = axes[i // n_cols, i % n_cols]
        # 准备 + 执行 时段
        T_win = T_AIM if c["type"] == '射击' else T_FOCUS
        i_start = c["idx_exec"] - int(T_win / kin["dt"])
        i_end = c["idx_exec"]
        idx_range = np.arange(max(i_start, 0), i_end + 1)
        t_w = kin["t"][idx_range]
        v_w = kin["speed"][idx_range]
        a_w = kin["accel"][idx_range]
        ax.plot(t_w, v_w, 'b-', lw=1, label='|v|')
        ax.plot(t_w, a_w, 'r-', lw=1, label='|a|')
        v_max = V_SHOOT_MAX if c["type"] == '射击' else V_PHOTO_MAX
        ax.axhline(v_max, color='orange', ls=':', lw=0.8)
        ax.axhline(A_SHOOT_MAX, color='red', ls=':', lw=0.8)
        ax.set_title(f"{c['target']} {c['type']} d={c['d']:.1f}", fontsize=8)
        ax.tick_params(labelsize=7)
        if i == 0:
            ax.legend(fontsize=7)
        ax.grid(alpha=0.3)
    # 隐藏多余子图
    for j in range(n_sel, n_rows * n_cols):
        axes[j // n_cols, j % n_cols].axis('off')
    plt.tight_layout()
    plt.savefig(FIGS / 'Q4_kinematic_profile.png', dpi=200)
    plt.close()


# =====================================================================
# Main
# =====================================================================
def main():
    print("=" * 72)
    print("Q4: 基于附件3 轨迹的多目标任务调度优化")
    print("=" * 72)

    kin = load_kinematic()
    S, S_id, P, P_id = load_targets()
    print(f"轨迹: {len(kin['t'])} 点, 时间 [{kin['t'][0]:.2f}, {kin['t'][-1]:.2f}] s, dt={kin['dt']:.3f}")
    print(f"目标: 射击 {len(S)} 个, 拍照 {len(P)} 个")

    # 阈值打印
    W_aim_s = int(round(T_AIM / kin["dt"]))
    W_aim_p = int(round(T_FOCUS / kin["dt"]))
    print(f"\n[约束 - 严格按题面]")
    print(f"  射击: d ∈ [{D_SHOOT_MIN}, {D_SHOOT_MAX}] m, v ≤ {V_SHOOT_MAX}, "
          f"|a| ≤ {A_SHOOT_MAX}, T_aim={T_AIM}s ({W_aim_s} 点), 命中率 {HIT_RATE}")
    print(f"  拍照: d ∈ [{D_PHOTO_MIN}, {D_PHOTO_MAX}] m, v ≤ {V_PHOTO_MAX}, "
          f"|a| ≤ {A_PHOTO_MAX}, T_focus={T_FOCUS}s ({W_aim_p} 点), 角度 ≥ {ANG_DIFF_DEG}°")

    # ===== 标称解 (P0 修正前的基线) =====
    print(f"\n========== 标称解 (无鲁棒约束, 无过渡时间 ε=0) ==========")
    shoot_cand_nom = shoot_candidates(S, S_id, kin, W_aim_s, robust=False)
    photo_cand_nom = photo_candidates(P, P_id, kin, W_aim_p,
                                      np.deg2rad(ANG_DIFF_DEG), robust=False)
    all_cand_nom = shoot_cand_nom + photo_cand_nom
    selected_nom = weighted_interval_scheduling_with_uniqueness(
        all_cand_nom, epsilon=0.0)
    selected_nom.sort(key=lambda c: c["t_prep"])
    n_shoot_nom = sum(1 for c in selected_nom if c["type"] == '射击')
    n_photo_nom = sum(1 for c in selected_nom if c["type"] == '拍照')
    print(f"  候选 {len(all_cand_nom)} → 入选 {len(selected_nom)} "
          f"({n_shoot_nom} 射击 + {n_photo_nom} 拍照)")

    # ===== 鲁棒解 (chance constraint + ε = 0.1s) =====
    print(f"\n========== 鲁棒解 (chance constraint z={ROBUST_Z}, ε={EPSILON_TRANSITION}s) ==========")
    print(f"  σ_x ≈ σ_y mean = {kin['sigma_x'].mean():.3f} m (KF 后验)")
    print(f"  σ_v = {SIGMA_V_REL*100:.0f}% × |v|")
    shoot_cand_rob = shoot_candidates(S, S_id, kin, W_aim_s, robust=True)
    photo_cand_rob = photo_candidates(P, P_id, kin, W_aim_p,
                                      np.deg2rad(ANG_DIFF_DEG), robust=True)
    all_cand_rob = shoot_cand_rob + photo_cand_rob
    selected_rob = weighted_interval_scheduling_with_uniqueness(
        all_cand_rob, epsilon=EPSILON_TRANSITION)
    selected_rob.sort(key=lambda c: c["t_prep"])
    n_shoot_rob = sum(1 for c in selected_rob if c["type"] == '射击')
    n_photo_rob = sum(1 for c in selected_rob if c["type"] == '拍照')
    print(f"  鲁棒候选 {len(all_cand_rob)} → 入选 {len(selected_rob)} "
          f"({n_shoot_rob} 射击 + {n_photo_rob} 拍照)")
    n_unique_shoot_rob = len({c["target"] for c in shoot_cand_rob})
    n_unique_photo_rob = len({c["target"] for c in photo_cand_rob})
    print(f"  鲁棒候选覆盖: 射击 {n_unique_shoot_rob}/{len(S)} 目标, "
          f"拍照 {n_unique_photo_rob}/{len(P)} 目标")

    # ===== 标称 vs 鲁棒 对照 =====
    print(f"\n========== 标称 vs 鲁棒 对照 ==========")
    nom_targets = {(c["target"], c["type"], round(c["t_exec"], 2)) for c in selected_nom}
    rob_targets = {(c["target"], c["type"], round(c["t_exec"], 2)) for c in selected_rob}
    only_nom = nom_targets - rob_targets
    only_rob = rob_targets - nom_targets
    common = nom_targets & rob_targets
    print(f"  公共: {len(common)} 任务")
    print(f"  仅标称 (鲁棒模式被剔除): {len(only_nom)}")
    for t in sorted(only_nom, key=lambda x: x[2]):
        print(f"    {t[0]} {t[1]} @ t={t[2]}")
    print(f"  仅鲁棒 (标称中不存在): {len(only_rob)}")
    for t in sorted(only_rob, key=lambda x: x[2]):
        print(f"    {t[0]} {t[1]} @ t={t[2]}")

    # 主交付: 鲁棒解 (P0 修正)
    print(f"\n========== 主交付: 鲁棒解 ==========")
    selected = selected_rob
    all_cand = all_cand_rob
    shoot_cand = shoot_cand_rob
    photo_cand = photo_cand_rob
    n_unique_shoot = n_unique_shoot_rob
    n_unique_photo = n_unique_photo_rob

    n_shoot_sel = sum(1 for c in selected if c["type"] == '射击')
    n_photo_sel = sum(1 for c in selected if c["type"] == '拍照')
    expected_hits = n_shoot_sel * HIT_RATE
    print(f"\n[调度结果] 候选总数 {len(all_cand)} → 入选 {len(selected)}")
    print(f"  射击 {n_shoot_sel} 次 (期望命中 {expected_hits:.2f})")
    print(f"  拍照 {n_photo_sel} 次")
    print(f"\n  最终方案:")
    print(f"  {'#':>3s} {'目标':>5s} {'任务':>5s} {'准备':>8s} {'执行':>8s} "
          f"{'d':>6s} {'v':>6s} {'a':>6s}")
    print("  " + "-" * 60)
    for i, c in enumerate(selected):
        print(f"  {i+1:>3d} {c['target']:>5s} {c['type']:>5s} "
              f"{c['t_prep']:>8.2f} {c['t_exec']:>8.2f} "
              f"{c['d']:>6.2f} {c['v']:>6.3f} {c['a']:>6.3f}")

    # 输出
    out_xlsx = write_result_xlsx(selected)
    print(f"\n[输出] {out_xlsx}")

    # 全部候选 + 入选标记
    df_cand = pd.DataFrame([{
        '目标': c['target'],
        '任务': c['type'],
        '开始准备时刻(s)': round(c['t_prep'], 3),
        '任务执行时刻(s)': round(c['t_exec'], 3),
        '距离(m)': round(c['d'], 3),
        '线速度(m/s)': round(c['v'], 3),
        '加速度(m/s²)': round(c['a'], 3),
        '视角(°)': round(np.rad2deg(c['view_angle']), 1) if c['type'] == '拍照' else None,
        '入选': '是' if c in selected else '否',
    } for c in all_cand])
    df_cand.to_excel(OUT / "Q4_全部候选.xlsx", index=False)
    print(f"[输出] {OUT / 'Q4_全部候选.xlsx'}")

    # 最终方案
    df_sel = pd.DataFrame([{
        '序号': i + 1, '目标编号': c['target'], '任务': c['type'],
        '开始准备时刻(s)': round(c['t_prep'], 2),
        '任务执行时刻(s)': round(c['t_exec'], 2),
        '距离(m)': round(c['d'], 3),
        '线速度(m/s)': round(c['v'], 3),
        '加速度(m/s²)': round(c['a'], 3),
    } for i, c in enumerate(selected)])
    df_sel.to_excel(OUT / "Q4_最终方案.xlsx", index=False)
    print(f"[输出] {OUT / 'Q4_最终方案.xlsx'}")

    # JSON 汇总: 标称 + 鲁棒 双解
    summary = dict(
        constraints=dict(
            shoot=dict(d_min=D_SHOOT_MIN, d_max=D_SHOOT_MAX,
                       v_max=V_SHOOT_MAX, a_max=A_SHOOT_MAX,
                       T_aim=T_AIM, hit_rate=HIT_RATE),
            photo=dict(d_min=D_PHOTO_MIN, d_max=D_PHOTO_MAX,
                       v_max=V_PHOTO_MAX, a_max=A_PHOTO_MAX,
                       T_focus=T_FOCUS, angle_diff_deg=ANG_DIFF_DEG),
            robustness=dict(
                z_value=ROBUST_Z,
                z_meaning="90% one-sided (chance constraint)",
                sigma_v_relative=SIGMA_V_REL,
                sigma_x_kf_mean=float(kin["sigma_x"].mean()),
                epsilon_transition_s=EPSILON_TRANSITION,
            ),
        ),
        n_targets=dict(shoot=len(S), photo=len(P)),
        nominal_solution=dict(
            n_candidates=len(all_cand_nom),
            n_selected=len(selected_nom),
            n_shoot=n_shoot_nom, n_photo=n_photo_nom,
            expected_hits=n_shoot_nom * HIT_RATE,
            note="标称解, 无鲁棒约束, 无过渡时间. 含边界值候选 (S06 d=5.02, P10 v=1.500).",
        ),
        robust_solution=dict(
            n_candidates=len(all_cand_rob),
            n_selected=len(selected_rob),
            n_shoot=n_shoot_rob, n_photo=n_photo_rob,
            expected_hits=n_shoot_rob * HIT_RATE,
            note=f"主交付. chance constraint z={ROBUST_Z} + ε={EPSILON_TRANSITION}s 过渡时间.",
        ),
        comparison=dict(
            common=len(common),
            only_nominal=len(only_nom),
            only_robust=len(only_rob),
            removed_by_robust=[{"target": t[0], "type": t[1], "t_exec": t[2]}
                               for t in sorted(only_nom, key=lambda x: x[2])],
            new_in_robust=[{"target": t[0], "type": t[1], "t_exec": t[2]}
                           for t in sorted(only_rob, key=lambda x: x[2])],
        ),
        n_candidates=dict(shoot=len(shoot_cand), photo=len(photo_cand)),
        n_unique_targets_with_candidates=dict(
            shoot=n_unique_shoot, photo=n_unique_photo),
        n_selected=dict(total=len(selected), shoot=n_shoot_sel, photo=n_photo_sel),
        expected_shoot_hits=expected_hits,
        objective_function_semantics=(
            "max ∑ x_i, 完成数 = 不同目标数 (而非射击次数). "
            "假设单次 0.85 命中已为工程容差, 无需重射. "
            "(三审一致建议, 见 Q4_三审综合.md §3)"),
        selected=[
            {
                "rank": i + 1,
                "target": c['target'],
                "type": c['type'],
                "t_prep": round(c['t_prep'], 3),
                "t_exec": round(c['t_exec'], 3),
                "distance_m": round(c['d'], 3),
                "speed_mps": round(c['v'], 3),
                "accel_mps2": round(c['a'], 3),
                "view_angle_deg": round(np.rad2deg(c['view_angle']), 1)
                                   if c['type'] == '拍照' else None,
            }
            for i, c in enumerate(selected)
        ],
    )
    (OUT / "Q4_summary.json").write_text(
        json.dumps(summary, indent=2, ensure_ascii=False),
        encoding="utf-8")
    print(f"[输出] {OUT / 'Q4_summary.json'}")

    # 出图
    plot_overview(kin, S, S_id, P, P_id, all_cand, selected)
    plot_gantt(selected)
    plot_kinematic_profile(kin, selected)
    print(f"[图] Q4_overview.png, Q4_gantt.png, Q4_kinematic_profile.png")


if __name__ == '__main__':
    main()
